from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook
import os
import requests
import datetime
import asyncio
import logging
import traceback
import tempfile
import vk_api
import pandas as pd
import re
from bs4 import BeautifulSoup
from typing import List, Dict, Optional

from dbrequests import delete_outdated_schedules, update_schedule

# Настройка логирования в stdout
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

time_from_pair = {
    "1 пара": "8:20-9:50",
    "2 пара": "10:00-11:30",
    "3 пара": "11:45-13:15",
    "4 пара": "14:00-15:30",
    "5 пара": "15:45-17:15",
    "6 пара": "17:20-18:50",
    "7 пара": "18:55-20:15",
}

user_agent = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36"
}
url_site = "https://www.vyatsu.ru/studentu-1/spravochnaya-informatsiya/zanyatost-auditoriy.html"
url_teacher_site = "https://www.vyatsu.ru/studentu-1/spravochnaya-informatsiya/teacher.html"

VK_TOKEN = "vk1.a.AjwujFeUGs7HScVf-EMqzP9q3_v0HqkhQCGt14fj2Jz05hmjGdkzk77JOGghcMMrNlkWjdc0rel4JDmgtf-DQKU2tQXtkWzmhHuymbeWDe05oNSdyfrXBER7jXcEUoWBGC7L6C65EXzKpKjljbOyYWU2EH8cq6mXhMTjm4JFNGqFUklkihhIQpRhI5Vd_I86lJ-P19ps-ODY-m8sJTAPqA"
GROUP_ID = "-85060840"

async def get_content(url):
    logger.info(f"Запрос контента с {url}")
    response = requests.get(url, headers=user_agent)
    response.raise_for_status()
    return response

async def get_urls(response):
    list_urls = []
    text: str = response.text

    while True:
        index = text.find('href="/reports/')
        if index == -1:
            break
        start_index = index + len('href="')
        end_index = text.find('"', start_index)
        url = "https://www.vyatsu.ru" + text[start_index:end_index]

        if url.endswith(".xls"):
            try:
                date_str = url[-12:-4]
                last_date = datetime.date(
                    year=int(date_str[-4:]),
                    month=int(date_str[-6:-4]),
                    day=int(date_str[-8:-6]),
                )
                current_date = datetime.date.today()
                if current_date < last_date and (last_date - current_date).days < 180:
                    list_urls.append(url)
                    logger.info(f"Найдена актуальная ссылка: {url}")
            except Exception as e:
                logger.error(f"Ошибка обработки ссылки {url}: {e}")
                logger.error(traceback.format_exc())
        text = text[end_index:]
    return list_urls

from bs4 import BeautifulSoup
import re
import datetime
import logging

logger = logging.getLogger(__name__)

def improved_parse_cell(cell_text: str) -> List[Dict[str, Optional[str]]]:
    if not cell_text or not isinstance(cell_text, str):
        return []

    entries = cell_text.strip().split("\n")
    results = []

    current = {
        "discipline": "",
        "lesson_type": "",
        "teacher": "",
        "cabinet": "",
        "subgroup": ""
    }

    for line in entries:
        line = line.strip()
        if not line:
            continue

        # Аудитория
        if re.match(r"^\d{1,2}-\d{1,3}$", line):
            current["cabinet"] = line
            continue

        # Преподаватель
        if re.match(r".+\s[А-ЯЁ]\.[А-ЯЁ]\.", line):
            current["teacher"] = line
            continue

        # Вид занятия
        if "занятие" in line or "Лекция" in line or "урок" in line:
            current["lesson_type"] = line
            continue

        # Подгруппа
        if "подгруппа" in line:
            current["subgroup"] = line
            continue

        # Всё остальное — дисциплина
        current["discipline"] += (line + " ")

    current = {k: v.strip() for k, v in current.items()}
    results.append(current)

    return results

async def get_teacher_urls(response):
    list_urls = []
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # Находим все кафедры на странице
    kafedry = soup.find_all('div', class_='kafPeriod')
    
    for kaf in kafedry:
        kafedra_name = kaf.text.strip()
        
        # Родительский блок, где лежат ссылки на расписание кафедры
        parent = kaf.parent
        
        # Ищем все ссылки на xls внутри этого блока
        links = parent.find_all('a', href=re.compile(r'/reports/schedule/prepod/.*\.xls'))
        
        for link in links:
            url = "https://www.vyatsu.ru" + link['href']
            teacher_name = link.text.strip()  # Извлекаем имя преподавателя из текста ссылки
            
            try:
                # Парсим дату начала из имени файла для фильтрации
                parts = url.split('_')
                if len(parts) >= 3:
                    start_date_str = parts[-2]
                    start_date = datetime.date(
                        year=int(start_date_str[4:8]),
                        month=int(start_date_str[2:4]),
                        day=int(start_date_str[0:2]),
                    )
                    current_date = datetime.date.today()
                    # Отбираем только актуальные файлы
                    if current_date < start_date and (start_date - current_date).days < 180:
                        list_urls.append((url, teacher_name, kafedra_name))
                        logger.info(f"Найдена ссылка кафедры: {url}, преподаватель: {teacher_name}, кафедра: {kafedra_name}")
            except Exception as e:
                logger.error(f"Ошибка обработки ссылки {url}: {e}")
    
    return list_urls



async def download(url: str) -> str:
    response = requests.get(url, headers=user_agent)
    response.raise_for_status()
    filename = url[-25:]
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xls") as temp_file:
        temp_file.write(response.content)
        temp_path = temp_file.name
    logger.info(f"Скачан файл {filename} в {temp_path}")
    return temp_path

async def convert_xls_to_xlsx(path: str) -> str:
    x2x = XLS2XLSX(path)
    xlsx_path = path + "x"
    x2x.to_xlsx(xlsx_path)
    logger.info(f"Конвертирован {path} в {xlsx_path}")
    return xlsx_path

async def parsing_url(url: str) -> None:
    try:
        path = await download(url)
        try:
            xlsx_path = await convert_xls_to_xlsx(path)
            os.remove(path)
            logger.info(f"Удален временный файл {path}")

            workbook = load_workbook(xlsx_path)
            worksheet = workbook.active

            for i in range(3, worksheet.max_column):
                cabinet_number: str = worksheet.cell(row=2, column=i).value
                if not cabinet_number:
                    continue

                for j in range(3, worksheet.max_row):
                    time_lesson: str = worksheet.cell(row=j, column=2).value
                    if not time_lesson:
                        continue
                    time_lesson = time_lesson.strip()
                    if time_lesson not in time_from_pair:
                        logger.warning(f"Неизвестное время пары: {time_lesson}")
                        continue
                    time_lesson = time_from_pair[time_lesson]

                    date = worksheet.cell(row=j, column=1).value
                    index_for_day = j - 1
                    while date is None:
                        date = worksheet.cell(row=index_for_day, column=1).value
                        index_for_day -= 1
                    date = date.strip()[-8:]
                    date_obj = datetime.datetime.strptime(date, "%d.%m.%y")
                    date = date_obj.strftime("%Y-%m-%d")

                    value: str = worksheet.cell(row=j, column=i).value
                    if value and "Резервирование" in value:
                        value = None

                    if value:
                        list_value = value.split("\n") if "\n" in value else [value]
                        list_value = [v.strip() for v in list_value]

                        for ii in range(len(list_value)):
                            if len(list_value[ii].split()[0]) <= 3:
                                list_value[ii] = " ".join(list_value[ii].split()[1:])

                        name_of_group = []
                        name_teacher = []
                        name_of_discipline = []

                        for value in list_value:
                            text_split = value.split()
                            if text_split[-1].count(".") == 2:
                                name_teacher.append(" ".join(text_split[-2:]))
                                text_split = text_split[:-2]
                            else:
                                name_teacher.append(None)

                            if text_split[0][-1] == ",":
                                name_of_group.append(" ".join(text_split[:3]))
                                text_split = text_split[3:]
                            else:
                                name_of_group.append(text_split[0])
                                text_split = text_split[1:]

                            name_of_discipline.append(" ".join(text_split))

                        if len(list_value) == 1:
                            await update_schedule(
                                date,
                                time_lesson,
                                cabinet_number,
                                name_of_group,
                                name_teacher,
                                name_of_discipline,
                            )
                        else:
                            await update_schedule(
                                date,
                                time_lesson,
                                cabinet_number,
                                name_of_group,
                                name_teacher,
                                name_of_discipline,
                                many=True,
                            )
                    else:
                        await update_schedule(
                            date, time_lesson, cabinet_number, None, None, None, empty=True
                        )

            os.remove(xlsx_path)
            logger.info(f"Удален файл {xlsx_path}")
        except Exception as e:
            logger.error(f"Ошибка парсинга файла {url}: {e}")
            logger.error(traceback.format_exc())
        finally:
            if os.path.exists(path):
                os.remove(path)
                logger.info(f"Удален временный файл {path}")
    except Exception as e:
        logger.error(f"Ошибка обработки URL {url}: {e}")
        logger.error(traceback.format_exc())

async def parsing_teacher_url(url: str, department: str) -> None:
    try:
        path = await download(url)
        try:
            xlsx_path = await convert_xls_to_xlsx(path)
            os.remove(path)
            logger.info(f"Удален временный файл {path}")

            workbook = load_workbook(xlsx_path)
            worksheet = workbook.active

            for row in range(2, worksheet.max_row + 1):
                date = worksheet.cell(row=row, column=1).value
                time_lesson = worksheet.cell(row=row, column=2).value  # Время теперь в колонке 2
                teacher_name = worksheet.cell(row=row, column=3).value  # Преподаватель в колонке 3
                combined_info = worksheet.cell(row=row, column=4).value  # Дисциплина и группа в колонке 4
                cabinet_number = worksheet.cell(row=row, column=5).value  # Аудитория в колонке 5

                if not all([date, time_lesson, teacher_name, combined_info, cabinet_number]):
                    continue

                # Приводим к строкам
                teacher_name = str(teacher_name).strip()
                if len(teacher_name) > 255:
                    logger.warning(f"Обрезано имя преподавателя: {teacher_name[:255]}...")
                    teacher_name = teacher_name[:255]

                date = str(date).strip()[-8:]
                try:
                    date_obj = datetime.datetime.strptime(date, "%d.%m.%Y")
                    date = date_obj.strftime("%Y-%m-%d")
                except ValueError as e:
                    logger.warning(f"Некорректный формат даты в строке {row}: {date}, пропуск")
                    continue

                time_lesson = str(time_lesson).strip()
                if time_lesson not in time_from_pair:
                    logger.warning(f"Неизвестное время пары в строке {row}: {time_lesson}, пропуск")
                    continue
                time_lesson = time_from_pair[time_lesson]

                # Разбираем combined_info на группу и дисциплину
                combined_info = str(combined_info).strip()
                name_group = "Unknown"
                name_discipline = combined_info
                if "кафедра" in combined_info.lower():
                    parts = combined_info.split("кафедра")
                    name_discipline = parts[0].strip()
                    if len(parts) > 1:
                        # Извлекаем группу, если она есть (например, "ИГЭ-171-23-01 доцент")
                        discipline_part = parts[0].strip()
                        group_match = re.search(r'([А-Яа-я0-9-]+-\d+-\d+-\d+)', discipline_part)
                        if group_match:
                            name_group = group_match.group(0)
                            name_discipline = discipline_part.replace(name_group, "").strip()
                elif re.search(r'([А-Яа-я0-9-]+-\d+-\d+-\d+)', combined_info):
                    group_match = re.search(r'([А-Яа-я0-9-]+-\d+-\d+-\d+)', combined_info)
                    name_group = group_match.group(0)
                    name_discipline = combined_info.replace(name_group, "").strip()

                if len(name_group) > 255:
                    logger.warning(f"Обрезано название группы: {name_group[:255]}...")
                    name_group = name_group[:255]
                if len(name_discipline) > 255:
                    logger.warning(f"Обрезано название дисциплины: {name_discipline[:255]}...")
                    name_discipline = name_discipline[:255]
                cabinet_number = str(cabinet_number).strip()
                if len(cabinet_number) > 50:
                    logger.warning(f"Обрезано название аудитории: {cabinet_number[:50]}...")
                    cabinet_number = cabinet_number[:50]
                if len(department) > 255:
                    logger.warning(f"Обрезано название кафедры: {department[:255]}...")
                    department = department[:255]

                logger.info(f"Запись в базу: date={date}, time_lesson={time_lesson}, cabinet_number={cabinet_number}, group={name_group}, teacher={teacher_name}, discipline={name_discipline}, department={department}")
                await update_schedule(
                    date,
                    time_lesson,
                    cabinet_number,
                    [name_group],
                    [teacher_name],
                    [name_discipline],
                    department=department
                )

            os.remove(xlsx_path)
            logger.info(f"Удален файл {xlsx_path}")
        except Exception as e:
            logger.error(f"Ошибка парсинга файла преподавателя {url}: {e}")
            logger.error(traceback.format_exc())
        finally:
            if os.path.exists(path):
                os.remove(path)
                logger.info(f"Удален временный файл {path}")
    except Exception as e:
        logger.error(f"Ошибка обработки URL преподавателя {url}: {e}")
        logger.error(traceback.format_exc())

async def download_vk_file(url, filename):
    response = requests.get(url)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(response.content)
        temp_path = temp_file.name
    logger.info(f"Скачан файл {filename} в {temp_path}")
    return temp_path

async def parse_vk_schedule_async():
    try:
        vk_session = vk_api.VkApi(token=VK_TOKEN, api_version="5.131")
        vk = vk_session.get_api()

        posts = vk.wall.get(owner_id=GROUP_ID, count=50, filter="all")
        items = posts.get("items", [])

        for post in items:
            if "расписание" not in post.get("text", "").lower():
                continue

            post_date = datetime.datetime.fromtimestamp(post["date"]).strftime("%Y-%m-%d %H:%M:%S")
            logger.info(f"Обработка поста от {post_date}, ID: {post['id']}")

            attachments = post.get("attachments", [])
            if not attachments:
                logger.info("В этом посте нет прикреплённых файлов.")
                continue

            for attachment in attachments:
                if attachment["type"] == "doc":
                    doc = attachment["doc"]
                    if doc["ext"] == "xlsx":
                        file_name = doc["title"]
                        file_url = doc["url"]

                        logger.info(f"Найден файл: {file_name}")
                        downloaded_file = await download_vk_file(file_url, file_name)

                        try:
                            schedules = await parse_schedule_structured(downloaded_file, file_name)
                            logger.info(f"Обработано {len(schedules)} записей из {file_name}")

                            # Временный лог для отладки
                            for entry in schedules[:5]:  # Показываем только первые 5 записей
                                logger.info(
                                    f"[ПРОВЕРКА ВК РАСПИСАНИЯ] {entry['date']} | {entry['time_lesson']} | "
                                    f"{entry['name_group']} | {entry['name_discipline']} | "
                                    f"{entry['name_teacher']} | {entry['cabinet_number']}"
    )


                            for entry in schedules:
                                await update_schedule(
                                    entry["date"],
                                    entry["time_lesson"],
                                    entry["cabinet_number"],
                                    [entry["name_group"]],
                                    [entry["name_teacher"]],
                                    [entry["name_discipline"]],
                                )

                            logger.info(f"РАСПИСАНИЕ из {file_name} сохранено в базу данных!")
                        finally:
                            if os.path.exists(downloaded_file):
                                os.remove(downloaded_file)
                                logger.info(f"Удален временный файл {downloaded_file}")
                    else:
                        logger.info(f"Файл {doc['title']} — не .xlsx, пропущен.")
                else:
                    logger.info("Не документ, пропущен.")
    except Exception as e:
        logger.error(f"Ошибка при парсинге VK: {e}")
        logger.error(traceback.format_exc())

async def parse_schedule_structured(file_path, file_name):
    try:
        date_match = re.search(r'(\d{2}\.\d{2})-(\d{2}\.\d{2})', file_name)
        if date_match:
            start_date_str, end_date_str = date_match.groups()
            start_date = datetime.datetime.strptime(f"{start_date_str}.2025", "%d.%m.%Y").date()
            end_date = datetime.datetime.strptime(f"{end_date_str}.2025", "%d.%m.%Y").date()
        else:
            raise ValueError(f"Не удалось извлечь диапазон дат из {file_name}")

        current_date = datetime.datetime.now().date()
        if start_date < current_date:
            start_date = current_date

        with pd.ExcelFile(file_path) as xl:
            schedules = []

            for sheet_name in xl.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

                header_row = None
                for idx, row in df.iterrows():
                    if any(isinstance(val, str) and 'День недели' in val for val in row):
                        header_row = idx
                        break

                if header_row is None:
                    logger.info(f"Пропуск листа {sheet_name}: не найдена строка 'День недели'")
                    continue

                day_cols = []
                time_cols = []
                for col in df.columns:
                    val = df.iloc[header_row, col]
                    if isinstance(val, str) and 'День недели' in val:
                        day_cols.append(col)
                        time_cols.append(col + 1)

                if not day_cols:
                    logger.info(f"Пропуск листа {sheet_name}: не найдены столбцы дней")
                    continue

                for block_idx, day_col in enumerate(day_cols):
                    time_col = time_cols[block_idx]
                    group_start_col = day_col + 2
                    group_cols = [(group_start_col + i * 4) for i in range(3)]
                    group_cols = [col for col in group_cols if col < df.shape[1]]
                    if not group_cols:
                        continue

                    group_names = []
                    for col in group_cols:
                        group_val = df.iloc[header_row - 1, col]
                        if isinstance(group_val, str) and 'Группа' in group_val:
                            group_name = group_val.replace("Группа", "").strip()
                        else:
                            group_name = "Unknown"
                        group_names.append(group_name)

                    current_date = None
                    for idx in range(header_row + 1, df.shape[0]):
                        day_val = str(df.iloc[idx, day_col]).strip()
                        if day_val and 'nan' not in day_val.lower():
                            try:
                                date_match = re.search(r'(\d{1,2}\.\d{2})', day_val)
                                if date_match:
                                    date_str = date_match.group(1)
                                    current_date = datetime.datetime.strptime(f"{date_str}.2025", "%d.%m.%Y").date()
                            except ValueError:
                                current_date = None

                        if current_date and (current_date < start_date or current_date > end_date or current_date < current_date):
                            continue

                        time_val = str(df.iloc[idx, time_col]).strip()
                        if 'nan' in time_val.lower() or not time_val:
                            continue

                        if re.match(r'^\d{1,2}\.\d{2}-\d{1,2}\.\d{2}$', time_val):
                            pass
                        elif time_val in time_from_pair:
                            time_val = time_from_pair[time_val]
                        else:
                            logger.warning(f"Некорректный формат времени: {time_val}, пропуск")
                            continue

                        for group_idx, group_col in enumerate(group_cols):
                            cell_value = "\n".join([
                                str(df.iloc[idx, group_col]).strip(),
                                str(df.iloc[idx, group_col + 2]).strip(),
                                str(df.iloc[idx, group_col + 3]).strip()
                            ])
                            parsed_items = improved_parse_cell(cell_value)

                            for item in parsed_items:
                                if not all([item["discipline"], item["teacher"], item["cabinet"]]):
                                    continue
                                if not re.match(r'^\d+-\d+$', item["cabinet"]):
                                    logger.warning(f"Некорректный номер аудитории: {item['cabinet']}, пропуск")
                                    continue
                                schedules.append({
                                    "name_group": group_names[group_idx],
                                    "date": current_date.strftime("%Y-%m-%d"),
                                    "time_lesson": time_val,
                                    "name_discipline": item["discipline"],
                                    "name_teacher": item["teacher"],
                                    "cabinet_number": item["cabinet"]
                                })
                                

        return schedules

    except Exception as e:
        logger.error(f"Ошибка обработки файла {file_path}: {e}")
        return []

async def start_parsing():
    await delete_outdated_schedules()
    try:
      ##  logger.info("Начало парсинга расписания преподавателей")
      ##  response = await get_content(url_teacher_site)
      ##  teacher_urls = await get_teacher_urls(response)
      ##  for url, teacher_name, department in teacher_urls:  # Обновлено
      ##      await parsing_teacher_url(url, teacher_name, department)
       ## logger.info("Парсинг расписания преподавателей завершен")

        logger.info("Начало парсинга расписания колледжа из VK")
        await parse_vk_schedule_async()
        logger.info("Парсинг расписания колледжа из VK завершен")
    except Exception as e:
        logger.error(f"Ошибка парсинга: {e}")
        logger.error(traceback.format_exc())